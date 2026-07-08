#!/usr/bin/env python3
# 產生 ~/Downloads/實價登錄比對.xlsx：每社區一分頁，供人工核對網站行情資料
# 用法：python3 scripts/generate-verify-excel.py（需先跑過 generate-market-data.mjs 讓快取就位）
import csv, json, glob, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

config = json.load(open('scripts/market-config.json'))['communities']
BAD = re.compile('親友|特殊關係|債權|債務|毛胚|急買|急賣|瑕疵|受災')
rows, parks = [], {}
for d in sorted(glob.glob('node_modules/.cache/lvr/*/')):
    try:
        for p in csv.DictReader(open(d+'h_lvr_land_a_park.csv', encoding='utf-8-sig')):
            sn = p.get('編號')
            if not sn or sn == 'Serial number': continue
            e = parks.setdefault(sn, {'price':0,'sqm':0.0})
            e['price'] += int(p.get('車位價格') or 0); e['sqm'] += float(p.get('車位面積平方公尺') or 0)
    except FileNotFoundError: pass
    for r in csv.DictReader(open(d+'h_lvr_land_a.csv', encoding='utf-8-sig')):
        if r.get('鄉鎮市區') and '鄉鎮市區' not in r['鄉鎮市區']: rows.append(r)

wb = Workbook(); wb.remove(wb.active)
head = ['交易年月日','門牌','樓層','格局','總面積','車位坪','車位價(萬)','車位來源','總價(萬)','單價(萬/坪)','狀態','備註']
hfill = PatternFill('solid', fgColor='E8EDE4'); xfill = PatternFill('solid', fgColor='F5D5D5'); efill = PatternFill('solid', fgColor='FFF3D6')

for c in config:
    ws = wb.create_sheet(c['keyword'][:28]); ws.append(head)
    for cell in ws[1]: cell.font = Font(bold=True); cell.fill = hfill
    def addr_ok(addr):
        if c.get('exclude') and any(x in addr for x in c['exclude']): return False
        if c.get('numRange'):
            import unicodedata
            half = ''.join(unicodedata.normalize('NFKC', ch) for ch in addr)
            import re as _re
            m = _re.search(r'(\d+)(?:之\d+)?號', half)
            if not m: return False
            n = int(m.group(1))
            if n < c['numRange'][0] or n > c['numRange'][1]: return False
        return True
    deals = []
    for r in rows:
        if r['鄉鎮市區'] != c['district'] or c['road'] not in r['土地位置建物門牌']: continue
        if not addr_ok(r['土地位置建物門牌']): continue
        if not (r['交易標的'] or '').startswith('房地'): continue
        done = r['建築完成年月'] or ''
        if len(done) < 5 or int(done[:len(done)-4]) != c['year']: continue
        pk = '車位移轉總面積(平方公尺)' if '車位移轉總面積(平方公尺)' in r else '車位移轉總面積平方公尺'
        pj = parks.get(r.get('編號',''), {'price':0,'sqm':0})
        d2 = dict(r)
        d2['_pp'] = int(r['車位總價元'] or 0) or pj['price']
        d2['_ps'] = float(r.get(pk) or 0) or pj['sqm']
        m = re.search(r'車位(\d+)', r['交易筆棟數'] or '')
        d2['_pc'] = int(m.group(1)) if m else 0
        deals.append(d2)
    deals.sort(key=lambda r: r['交易年月日'], reverse=True)
    tiers = c.get('parkTiers')
    est_p = (c.get('parkWan',0)*10000) or next((d['_pp']/d['_pc'] for d in deals if d['_pp']>0 and d['_pc']>0), 0)
    est_s = next((d['_ps']/d['_pc'] for d in deals if d['_ps']>0 and d['_pc']>0), 20.0)
    src_label = '車位分級行情' if tiers else ('Cindy物件行情' if c.get('parkWan') else '社區最近成交')
    for r in deals:
        total = int(r['總價元'] or 0); sqm = float(r['建物移轉總面積平方公尺'] or 0)
        pp, ps, pc = r['_pp'], r['_ps'], r['_pc']
        note = r['備註'] or ''; bad = bool(BAD.search(note))
        has_park = pc>0 or ps>0 or pp>0
        src = ''
        split = pp>0 and ps>0
        if pp>0: src = '官方登記'
        hp = round(sqm*0.3025, 2); pping = round(ps*0.3025, 2)  # hp=總面積(比照官網)
        if bad: unit, status = '', '❌特殊交易(不上網站)'
        elif split:
            unit = round((total-pp)/((sqm-ps)*0.3025)/10000, 2)
            status = '✅公式二(已拆車位)'
        else:
            unit = round(total/(sqm*0.3025)/10000, 2) if sqm else ''
            status = '✅公式一' + ('(單價含車位🚗)' if has_park else '')
        layout = f"{r['建物現況格局-房']}房{r['建物現況格局-廳']}廳{r['建物現況格局-衛']}衛"
        dt = r['交易年月日']
        ws.append([f"{dt[:len(dt)-4]}/{int(dt[-4:-2])}/{int(dt[-2:])}", r['土地位置建物門牌'].replace('桃園市',''),
                   r['移轉層次'], layout, hp, pping or '', round(pp/10000,1) or '', src,
                   round(total/10000), unit, status, note[:50]])
        if bad:
            for cell in ws[ws.max_row]: cell.fill = xfill
        elif '含車位🚗' in status:
            for cell in ws[ws.max_row]: cell.fill = efill
    for i,w in enumerate([11,32,10,12,9,8,10,18,9,11,20,40],1):
        ws.column_dimensions[ws.cell(1,i).column_letter].width = w
    ws.freeze_panes = 'A2'
wb.save('/Users/wangweixun/Downloads/實價登錄比對.xlsx')
print('完成: ~/Downloads/實價登錄比對.xlsx')
